import os
import logging
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    Response,
    send_from_directory,
    jsonify,
)
import threading
import uuid
import csv
import openpyxl
from io import StringIO, BytesIO


## Python Modules
from .modules.db_write import (
    create_db,
    cleanup_incomplete,
    update_session_status,
)
from .modules.db_read import (
    get_logs,
    get_session_status,
    get_categories_from_db,
)
from parser_app.modules.parser_modules.core import parse_catalog, parse_product_urls
from .modules.parser_modules.scraper import scrape_categories
from .modules.logger import log_message
from .modules.utilities import (
    get_db_connection,
)


create_db()
app = Flask(__name__)
app.secret_key = os.urandom(24)
IMAGES_FOLDER = os.path.join(app.static_folder, "images")
CANCEL_FLAGS = {}  # Tracks cancellation per session_id


def export_to_csv(data, columns, filename):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in data:
        writer.writerow(row)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={filename}"},
    )


def export_to_xlsx(data, columns, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(columns)
    for row in data:
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"},
    )


@app.route("/", methods=["GET"])
def index():
    try:
        categories = scrape_categories()
        log_message(
            None,
            f"Rendering index.html with {len(categories)} categories",
            level="debug",
        )
        return render_template("index.html", parsing_status=None, categories=categories)
    except Exception as e:
        log_message(None, f"Error rendering index.html: {e}", level="error")
        return Response(f"Error rendering page: {e}", status=500)


@app.route("/favicon.ico")
def favicon():
    favicon_path = os.path.join(app.static_folder, "favicon.ico")
    if os.path.exists(favicon_path):
        return send_from_directory(app.static_folder, "favicon.ico")
    else:
        log_message(None, "favicon.ico not found in static folder", level="warning")
        return Response(status=404)


@app.route("/parse", methods=["POST"])
def parse():
    url = request.form.get("url", "").strip()
    category_url = request.form.get("category", "")
    category_name = request.form.get("category_name", "")
    max_pages = request.form.get("max_pages")
    max_products = request.form.get("max_products")

    max_pages = int(max_pages) if max_pages and max_pages.isdigit() else None
    max_products = (
        int(max_products) if max_products and max_products.isdigit() else None
    )

    session_id = str(uuid.uuid4())

    session["parse_session_id"] = session_id

    update_session_status(
        session_id,
        "in_progress",
        progress="collecting_urls",
        category_name=category_name,
    )
    CANCEL_FLAGS[session_id] = False

    log_message(
        session_id,
        f"URL: {url}, category_url: {category_url}, category_name:'{category_name}'",
        level="debug",
    )

    def run_parse(session_id, category_name):
        try:
            parse_catalog(
                catalog_url=category_url or None,
                category=category_name or None,
                max_pages=max_pages,
                max_products=max_products,
                session_id=session_id,
                url=url or None,
                cancel_flags=CANCEL_FLAGS,
            )

        except Exception as e:
            log_message(session_id, f"❌ Ошибка во время парсинга: {e}", level="error")
            update_session_status(session_id, "error")
            cleanup_incomplete(session_id)
        finally:
            CANCEL_FLAGS.pop(session_id, None)

    log_message(
        session_id,
        f"Starting parsing thread for session {session_id} (URL: {url}, Category: {category_name})",
        level="debug",
    )
    threading.Thread(target=run_parse, args=(session_id, category_name)).start()
    return redirect(url_for("results"))


@app.route("/confirm_parse", methods=["GET", "POST"])
def confirm_parse():
    parse_session_id = session.get("parse_session_id")
    if not parse_session_id:
        return redirect(url_for("index"))

    status, product_urls, progress = get_session_status(parse_session_id)
    if status != "awaiting_confirmation":
        return redirect(url_for("results"))

    if request.method == "POST":
        action = request.form.get("action")
        if action == "confirm":

            def continue_parse():
                try:
                    update_session_status(
                        parse_session_id,
                        "in_progress",
                        product_urls,
                        "parsing_products",
                    )
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    cursor.execute(
                        "SELECT category_name FROM parse_sessions WHERE session_id = ?",
                        (parse_session_id,),
                    )
                    category_name = cursor.fetchone()[0]
                    conn.close()
                    parse_product_urls(
                        product_urls, category_name, parse_session_id, CANCEL_FLAGS
                    )
                    if not CANCEL_FLAGS.get(parse_session_id, False):
                        update_session_status(parse_session_id, "complete")
                        log_message(
                            parse_session_id,
                            "✅ Парсинг завершен успешно.",
                            level="info",
                        )
                except Exception as e:
                    log_message(
                        parse_session_id,
                        f"❌ Ошибка во время парсинга: {e}",
                        level="error",
                    )
                    update_session_status(parse_session_id, "error")
                    cleanup_incomplete(parse_session_id)
                finally:
                    CANCEL_FLAGS.pop(parse_session_id, None)

            log_message(
                parse_session_id,
                "✅ Пользователь подтвердил продолжение парсинга",
                level="info",
            )

            threading.Thread(target=continue_parse).start()
            return redirect(url_for("results"))
        else:
            CANCEL_FLAGS[parse_session_id] = True
            update_session_status(parse_session_id, "canceled")
            cleanup_incomplete(parse_session_id)

            log_message(
                parse_session_id, "⚠️ Пользователь отменил парсинг", level="warning"
            )

            CANCEL_FLAGS.pop(parse_session_id, None)
            return redirect(url_for("results"))

    log_message(None, f"Function confirm_parse(), {product_urls}", level="debug")
    if len(product_urls):
        log_message(
            None, f"Function confirm_parse(), {len(product_urls)}", level="debug"
        )
    else:
        logging.error(f"len({product_urls})")
    return render_template(
        "confirm.html",
        product_count=len(product_urls),
        parse_session_id=parse_session_id,
    )


@app.route("/cancel_parse", methods=["POST"])
def cancel_parse():
    parse_session_id = session.get("parse_session_id")
    if parse_session_id:
        CANCEL_FLAGS[parse_session_id] = True
        update_session_status(parse_session_id, "canceled")
        cleanup_incomplete(parse_session_id)
        log_message(
            parse_session_id, "⚠️ Парсинг отменен пользователем", level="warning"
        )
        CANCEL_FLAGS.pop(parse_session_id, None)
    return redirect(url_for("results"))


@app.route("/status/<session_id>", methods=["GET"])
def status(session_id):
    status, _, _ = (
        get_session_status(session_id) if session_id else ("complete", [], None)
    )
    selected_category = request.args.get("category", None)
    if not selected_category and session_id:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT category_name FROM parse_sessions WHERE session_id = ?",
            (session_id,),
        )
        result = cursor.fetchone()
        selected_category = result[0] if result else None
        conn.close()

    conn = get_db_connection()
    cursor = conn.cursor()

    if selected_category and selected_category != "all":
        cursor.execute(
            "SELECT * FROM products WHERE category = ? AND is_complete = 1",
            (selected_category),
        )
    else:
        cursor.execute("SELECT * FROM products WHERE is_complete = 1")
    products = cursor.fetchall()

    cursor.execute("PRAGMA table_info(products)")
    product_columns = [col[1] for col in cursor.fetchall()]

    if selected_category and selected_category != "all":
        cursor.execute(
            """
            SELECT v.* FROM variants v
            JOIN products p ON v.product_id = p.id
            WHERE p.category = ? AND v.is_complete = 1
            """,
            (selected_category,),
        )
    else:
        cursor.execute("SELECT * FROM variants WHERE is_complete = 1")
    variants = cursor.fetchall()

    cursor.execute("PRAGMA table_info(variants)")
    variant_columns = [col[1] for col in cursor.fetchall()]

    logs = get_logs(session_id) if session_id else []

    conn.close()

    return jsonify(
        {
            "status": status,
            "logs": logs,
            "products": products,
            "product_columns": product_columns,
            "variants": variants,
            "variant_columns": variant_columns,
        }
    )


@app.route("/results")
def results():
    parse_session_id = session.get("parse_session_id", None)

    status, _, _ = (
        get_session_status(parse_session_id)
        if parse_session_id
        else ("complete", [], None)
    )

    selected_category = request.args.get("category", None)

    if not selected_category and parse_session_id:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT category_name FROM parse_sessions WHERE session_id = ?",
            (parse_session_id,),
        )
        result = cursor.fetchone()
        selected_category = result[0] if result else None
        conn.close()

    export_type = request.args.get("export", None)
    export_table = request.args.get("table", None)

    conn = get_db_connection()
    cursor = conn.cursor()

    categories = get_categories_from_db()

    if selected_category and selected_category != "all":
        cursor.execute(
            "SELECT * FROM products WHERE category = ? AND is_complete = 1",
            (selected_category,),
        )
    else:
        cursor.execute("SELECT * FROM products WHERE is_complete = 1")
    products = cursor.fetchall()

    cursor.execute("PRAGMA table_info(products)")
    product_columns = [col[1] for col in cursor.fetchall()]

    if selected_category and selected_category != "all":
        cursor.execute(
            """
            SELECT v.* FROM variants v
            JOIN products p ON v.product_id = p.id
            WHERE p.category = ? AND v.is_complete = 1
            """,
            (selected_category,),
        )
    else:
        cursor.execute("SELECT * FROM variants WHERE is_complete = 1")
    variants = cursor.fetchall()

    cursor.execute("PRAGMA table_info(variants)")
    variant_columns = [col[1] for col in cursor.fetchall()]

    logs = get_logs(parse_session_id) if parse_session_id else []

    if export_type and export_table:
        if export_table == "products":
            data = products
            columns = product_columns
            filename_prefix = "products"
        elif export_table == "variants":
            data = variants
            columns = variant_columns
            filename_prefix = "variants"
        else:
            data = []
            columns = []
            filename_prefix = "unknown"

        if export_type == "csv":
            return export_to_csv(data, columns, f"{filename_prefix}.csv")
        elif export_type == "xlsx":
            return export_to_xlsx(data, columns, f"{filename_prefix}.xlsx")

    conn.close()

    return render_template(
        "results.html",
        logs=logs,
        products=products,
        product_columns=product_columns,
        variants=variants,
        variant_columns=variant_columns,
        parsing_status=status,
        categories=categories,
        selected_category=selected_category,
        parse_session_id=parse_session_id,
    )


@app.route("/logs")
def get_latest_logs():
    parse_session_id = session.get("parse_session_id")
    logs = get_logs(parse_session_id) if parse_session_id else []
    return jsonify(logs)


@app.route("/parsing-status")
def get_parsing_status():
    parse_session_id = session.get("parse_session_id")
    status, _, _ = (
        get_session_status(parse_session_id)
        if parse_session_id
        else ("complete", [], None)
    )
    return jsonify({"status": status})


@app.route("/browse")
def browse():
    selected_category = request.args.get("category", None)

    conn = get_db_connection()
    cursor = conn.cursor()

    categories = get_categories_from_db()

    if selected_category and selected_category != "all":
        cursor.execute(
            "SELECT * FROM products WHERE category = ? AND is_complete = 1",
            (selected_category,),
        )
    else:
        cursor.execute("SELECT * FROM products WHERE is_complete = 1")
    products = cursor.fetchall()

    cursor.execute("PRAGMA table_info(products)")
    product_columns = [col[1] for col in cursor.fetchall()]

    product_data = []
    for product in products:
        product_id = product[0]
        cursor.execute(
            """
            SELECT * FROM variants WHERE product_id = ? AND is_complete = 1
            ORDER BY 
                CASE 
                    WHEN article_number GLOB '[0-9]*' THEN CAST(article_number AS REAL)
                    WHEN article_number GLOB '[0-9]*.[0-9]*' THEN CAST(article_number AS REAL)
                    ELSE article_number
                END
            """,
            (product_id,),
        )
        variants = cursor.fetchall()
        image_path = product[8]
        if image_path and os.path.exists(os.path.join(app.static_folder, image_path)):
            logging.debug(f"Image path verified: {image_path}")
        else:
            logging.warning(f"Image path missing or invalid: {image_path}")
        logging.debug(f"Product: {product[2]}, Variants: {len(variants)}")
        if variants:
            article_numbers = [v[2] for v in variants]
            logging.debug(f"Variant article_numbers (sorted): {article_numbers}")
            try:
                sorted_check = sorted(
                    article_numbers,
                    key=lambda x: float(x) if x.replace(".", "", 1).isdigit() else x,
                )
                if article_numbers == sorted_check:
                    logging.debug(
                        "Sorting verified: article_numbers are in ascending order"
                    )
                else:
                    logging.warning(
                        f"Sorting error: Expected {sorted_check}, got {article_numbers}"
                    )
            except ValueError as e:
                logging.error(f"Sorting verification failed: {e}")
        product_data.append({"product": product, "variants": variants})

    cursor.execute("PRAGMA table_info(variants)")
    variant_columns = [col[1] for col in cursor.fetchall()]

    parse_session_id = session.get("parse_session_id", None)
    status = get_session_status(parse_session_id)[0] if parse_session_id else None

    conn.close()

    return render_template(
        "browse.html",
        products=product_data,
        product_columns=product_columns,
        variant_columns=variant_columns,
        categories=categories,
        selected_category=selected_category,
        parsing_status=status,
    )


@app.route("/export_csv")
def export_csv():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM products WHERE is_complete = 1")
    products = cursor.fetchall()
    cursor.execute("PRAGMA table_info(products)")
    product_columns = [col[1] for col in cursor.fetchall()]
    conn.close()
    return export_to_csv(products, product_columns, "products.csv")


@app.route("/export_xlsx")
def export_xlsx():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM products WHERE is_complete = 1")
    products = cursor.fetchall()

    cursor.execute("PRAGMA table_info(products)")
    product_columns = [col[1] for col in cursor.fetchall()]

    cursor.execute("SELECT * FROM variants WHERE is_complete = 1")
    variants = cursor.fetchall()

    cursor.execute("PRAGMA table_info(variants)")
    variant_columns = [col[1] for col in cursor.fetchall()]

    conn.close()

    wb = openpyxl.Workbook()
    ws_products = wb.active
    ws_products.title = "Products"
    ws_products.append(product_columns)
    for row in products:
        ws_products.append(row)

    ws_variants = wb.create_sheet("Variants")
    ws_variants.append(variant_columns)
    for row in variants:
        ws_variants.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=products.xlsx"},
    )


if __name__ == "__main__":
    create_db()
    app.run(debug=True)
